from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import printDataStructure
import sys

class XLSX:

	def __init__(self, prazos, audiencias, acompanhamentos, plan_path="src/planilha.xlsx"):
		print("[XLSX] Initializing module", end=".")
		sys.stdout.flush()

		# Basic variables
		self.prazos = prazos
		self.audiencias = audiencias
		self.acompanhamentos = acompanhamentos
		self.plan_path = plan_path

		self.max_rows = 100 #Defoult. Maybe change at load_data func
		self.advogados = ["Cabral", "Vitória Tiannamen", "Raul Lobato", "Paulo Toledo"]
		print("", end=".")
		sys.stdout.flush()

		# Variables used to append secion indexes
		self.prazos_rows = list()	#Two indexes (init, end)
		self.audiencias_rows = list()	#Two indexes (init, end)
		self.acompanhamentos_rows = list()	#One index (init)
		print("", end=".")
		sys.stdout.flush()

		#Print structures
		self.print_prazos = printDataStructure.PrintDataStructure(1)
		self.print_audiencias_prazos = printDataStructure.PrintDataStructure(2)
		self.print_acompanhamento_prazos = printDataStructure.PrintDataStructure(3)
		print(" Done")
		sys.stdout.flush()

		# Elements
		self.load_and_write_data()

	def to_day_transform(self, data):
		d, m, a = data.split('/')
		d += m*30
		d += a*365
		return d

	def sort_dict_lists(self, dict_recived): #By date sorting
		for list_by_user in self.advogados:
			lista = dict_recived[list_by_user]
			lista.sort(key=lambda item: self.to_day_transform(item['Prazo']), reverse=False)	
			dict_recived[list_by_user] = lista
		return dict_recived


	def load_and_write_data(self):

		print("[XLSX] Loading content", end=".")
		sys.stdout.flush()

		#Types
		plan = list()
		tipo = ["CLIENTE", "ESTUDOS E ACOMPANHAMENTOS", "AUDIÊNCIAS E JULGAMENTOS", "PRAZOS"]
		self.estilos = [None for i in range(4)]

		################################################
		#	Style types
		self.width_dict = {'A': None, 'B': None, 'C': None, 'D': None, 'E': None, 'F': None, 'G': None, 'H': None, 'I': None}

		#0 -> PRAZOS LINE
		#1 -> CLIENTE LINE
		#2 -> NORMAL LINES
		#3 -> LABEL LINES
		self.styles = [list(), list(), list(), list()]

		wb = load_workbook(self.plan_path)
		ws = wb.worksheets[0]

		try:
			self.max_rows = ws.max_row
		except:
			print("[XLSX] Impossible to use dynamic rows into file. Using default 100 rows counter")

		for row in ws.iter_rows(min_row=1, max_col=9 , max_row=self.max_rows):
			aux = list()
			for cell in row:
				aux.append(cell.value)

			plan.append(aux)

		# Searching for column width
		for i in self.width_dict.keys():
			self.width_dict[i] = ws.column_dimensions[i].width

		print("", end=".")
		sys.stdout.flush()


		# searching for indexes
		# Will be used later in order to define how much lines
		# Each secion will use (Prazos, Audiencias and Acompanhamentos)
		index = 0
		for i in plan:
			index += 1
			for j in i:
				if tipo[0] == str(j).strip():
					self.prazos_rows.append(index+2)
				if tipo[1] == str(j).strip():
					self.prazos_rows.append(index-1)
					self.acompanhamentos_rows.append(index+2)
				if tipo[2] == str(j).strip():
					self.acompanhamentos_rows.append(index-1)
					self.audiencias_rows.append(index+2)
		print("", end=".")
		sys.stdout.flush()

		#Getting styles
		prazo_line = self.prazos_rows[0]-3
		cliente_line = self.prazos_rows[0]-2
		normal_line = self.prazos_rows[0]
		estudo_line = self.prazos_rows[1]+1
		for i in self.width_dict.keys():
			cel_prazo = str(i)+str(prazo_line)
			cel_cliente = str(i)+str(cliente_line)
			cel_normal = str(i)+str(normal_line)
			cel_estudo = str(i)+str(estudo_line)

			self.styles[0].append(ws[cel_prazo]._style)
			self.styles[1].append(ws[cel_cliente]._style)
			self.styles[2].append(ws[cel_normal]._style)
			self.styles[3].append(ws[cel_estudo]._style)

		# Adding to Data structures that will be used later
		# in order to build new xlsx
		index = 0
		for i in plan:
			if index >= self.prazos_rows[0]-1 and index < self.prazos_rows[1]:
				self.add_to_prazos(i)
			elif index >= self.acompanhamentos_rows[0]-1 and index < self.acompanhamentos_rows[1]:
				self.add_to_acompanhamentos(i)
			elif index >= self.audiencias_rows[0]-1:
				self.add_to_audiencias(i)
			index += 1
		print("", end=".")
		sys.stdout.flush()

		# Sorting data before assingment
		self.prazos = self.sort_dict_lists(self.prazos)
		#self.acompanhamentos = self.sort_dict_lists(self.acompanhamentos)
		#self.audiencias = self.sort_dict_lists(self.audicencias)
		print(" Done")

		# Indexing insto print_data_structure
		# For Prazos
		print("[XLSX] Creating new datasheet", end=".")
		for user in self.prazos.keys():
			for data in self.prazos[user]:
				if user == self.advogados[0]:
					self.print_prazos.add_to_cabral(data)
				elif user == self.advogados[1]:
					self.print_prazos.add_to_vitoria(data)
				elif user == self.advogados[2]:
					self.print_prazos.add_to_raul(data)
				elif user == self.advogados[3]:
					self.print_prazos.add_to_paulo(data)
		print("", end=".")
		sys.stdout.flush()

		# For Audiencias
		for user in self.audiencias.keys():
			for data in self.audiencias[user]:
				if user == self.advogados[0]:
					self.print_audiencias_prazos.add_to_cabral(data)
				elif user == self.advogados[1]:
					self.print_audiencias_prazos.add_to_vitoria(data)
				elif user == self.advogados[2]:
					self.print_audiencias_prazos.add_to_raul(data)
				elif user == self.advogados[3]:
					self.print_audiencias_prazos.add_to_paulo(data)
		print("", end=".")
		sys.stdout.flush()

		# For Acompanhamentos
		for user in self.acompanhamentos.keys():
			for data in self.acompanhamentos[user]:
				if user == self.advogados[0]:
					self.print_acompanhamento_prazos.add_to_cabral(data)
				elif user == self.advogados[1]:
					self.print_acompanhamento_prazos.add_to_vitoria(data)
				elif user == self.advogados[2]:
					self.print_acompanhamento_prazos.add_to_raul(data)
				elif user == self.advogados[3]:
					self.print_acompanhamento_prazos.add_to_paulo(data)

		# Agrouping at print_data_structe
		self.agroup()
		print("", end=".")
		sys.stdout.flush()

		#Gets structure to be written
		written_prazos = self.print_prazos.get_main_structure()
		written_audiencias = self.print_audiencias_prazos.get_main_structure()
		written_acompanhamentos = self.print_acompanhamento_prazos.get_main_structure()

		written_all = list()
		for i in written_prazos:
			written_all.append(i)
		for i in written_acompanhamentos:
			written_all.append(i)
		for i in written_audiencias:
			written_all.append(i)

		print("", end=".")
		sys.stdout.flush()

		# Creating new sheet in order to append new data
		ws = wb.create_sheet("Planilha atualizada", 0)

		# Dropping old sheed
		sheet_names = wb.get_sheet_names()
		std = wb.get_sheet_by_name(sheet_names[-1])
		wb.remove_sheet(std) #Droping old sheet

		for i in written_all:
			ws.append(i)

		#Loading new plan
		plan = list()
		try:
			self.max_rows = ws.max_row
		except:
			print("[XLSX] Impossible to use dynamic rows into file. Using default 100 rows counter")

		for row in ws.iter_rows(min_row=1, max_col=9 , max_row=self.max_rows):
			aux = list()
			for cell in row:
				aux.append(cell.value)
			plan.append(aux)

		# searching for new indexes
		# Will be used later in order to define how much lines
		# Each secion will use (Prazos, Audiencias and Acompanhamentos)
		index = 0
		self.prazos_rows = list()
		self.acompanhamentos_rows = list()
		self.audiencias_rows = list()
		for i in plan:
			index += 1
			for j in i:
				if tipo[0] == str(j).strip():
					self.prazos_rows.append(index+1)
				if tipo[1] == str(j).strip():
					self.prazos_rows.append(index-1)
					self.acompanhamentos_rows.append(index+1)
				if tipo[2] == str(j).strip():
					self.acompanhamentos_rows.append(index-1)
					self.audiencias_rows.append(index+1)

		print("", end=".")
		sys.stdout.flush()		

		#Creating new style
		for i in self.width_dict.keys():
			ws.column_dimensions[i].width = self.width_dict[i]

		#Setting name background collor
		yellowFill = PatternFill(start_color='FFFF00',
                   end_color='FFFF00',
                   fill_type='solid')

		#Setting styles
		index = 0
		for row in ws.iter_rows(min_row=1, max_col=9 , max_row=self.max_rows):
			index += 1
			it = 0
			for cel in row:
				try:

					flag_first = False
					if index == 2:
						cel._style = self.styles[0][it]
					if index == 3:
						cel._style = self.styles[1][it]
					if index >= self.prazos_rows[0] and index <= self.prazos_rows[1]:
						cel._style = self.styles[2][it]
					if index == self.prazos_rows[1]+1:
						cel._style = self.styles[3][it]
					if index >= self.acompanhamentos_rows[0] and index <= self.acompanhamentos_rows[1]:
						cel._style = self.styles[2][it]
					if index == self.acompanhamentos_rows[1]+1:
						cel._style = self.styles[3][it]
					if index >= self.audiencias_rows[0]:
						cel._style = self.styles[2][it]
				except:
					continue

				it +=1

		print(" Done")

		print("[XLSX] Writing file")

		#Editing name
		var = self.plan_path.split('/')

		# Escrevendo
		wb.save("src/planilha_updated.xlsx")

		print("[XLSX] Completed")

	def agroup(self):
		self.print_prazos.set_main_structure()
		self.print_audiencias_prazos.set_main_structure()
		self.print_acompanhamento_prazos.set_main_structure()

	#########################################
	##	Appends
	def add_to_prazos(self, lista):
		if lista[4] == None:
			return

		#Dict Structure
		data = {
			"processo":None,
			"descricao":None,
			"Tarefa":None,
			"Cliente":None,
			"Criado em":None,
			"Prazo Fatal":None,
			"Prazo":None,
			"Realizado em":None,
			"Advogado":None
		}

		data['processo'] = lista[3]
		data['descricao'] = lista[6]
		data['Tarefa'] = lista[5]
		data['Cliente'] = lista[2]
		try:
			data['Criado em'] = lista[1].strftime("%d/%m/%Y")
		except:
			data['Criado em'] = None
		try:
			data['Prazo Fatal'] = lista[7].strftime("%d/%m/%Y")
		except:
			data['Prazo Fatal'] = None
		try:
			data['Prazo'] = lista[8].strftime("%d/%m/%Y")
		except:
			data['Prazo'] = None
		data['Advogado'] = lista[4]

		self.prazos[lista[4]].append(data)

	def add_to_acompanhamentos(self, lista):
		if lista[4] == None:
			return

		#Dict Structure
		data = {
			"processo":None,
			"descricao":None,
			"Tarefa":None,
			"Cliente":None,
			"Criado em":None,
			"Prazo Fatal":None,
			"Prazo":None,
			"Realizado em":None,
			"Advogado":None
		}

		data['processo'] = lista[3]
		data['descricao'] = lista[6]
		data['Tarefa'] = lista[5]
		data['Cliente'] = lista[2]
		try:
			data['Criado em'] = lista[1].strftime("%d/%m/%Y")
		except:
			data['Criado em'] = None
		try:
			data['Prazo Fatal'] = lista[7].strftime("%d/%m/%Y")
		except:
			data['Prazo Fatal'] = None
		try:
			data['Prazo'] = lista[8].strftime("%d/%m/%Y")
		except:
			data['Prazo'] = None
		data['Advogado'] = lista[4]

		self.acompanhamentos[lista[4]].append(data)
		 #self.prazos[lista[4]].append(data)

	def add_to_audiencias(self, lista):
		if lista[4] == None:
			return

		#Dict Structure
		data = {
			"processo":None,
			"descricao":None,
			"Tarefa":None,
			"Cliente":None,
			"Criado em":None,
			"Prazo Fatal":None,
			"Prazo":None,
			"Realizado em":None,
			"Advogado":None
		}

		data['processo'] = lista[3]
		data['descricao'] = lista[6]
		data['Tarefa'] = lista[5]
		data['Cliente'] = lista[2]
		try:
			data['Criado em'] = lista[1].strftime("%d/%m/%Y")
		except:
			data['Criado em'] = None
		try:
			data['Prazo Fatal'] = lista[7].strftime("%d/%m/%Y")
		except:
			data['Prazo Fatal'] = None
		try:
			data['Prazo'] = lista[8].strftime("%d/%m/%Y")
		except:
			data['Prazo'] = None
		data['Advogado'] = lista[4]

		self.audiencias[lista[4]].append(data)
		#self.prazos[lista[4]].append(data)	